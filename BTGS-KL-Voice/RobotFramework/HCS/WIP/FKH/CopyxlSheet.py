from openpyxl import load_workbook
from copy import copy, deepcopy

class ExcelUtility(object):
	ROBOT_LIBRARY_SCOPE = 'Global'
	
	def __init__(self):
		print 'Read Cell Value in Excel File'
		
	def copy_excel_sheet(self,excelfile,sheetname):
		wb=load_workbook(filename=excelfile, data_only=True)
		#sheet_ranges=wb[sheetname]
		#sheet_ranges=wb.get_sheet_by_name(sheetname)
		#ws=wb.active
		for sheet in wb.worksheets:
			#return wb.sheetnames
			sheetTitle= sheet.title
			target= wb.copy_worksheet(sheet)
			wb.remove(sheet)
			#target.title= sheetTitle
		wb.active=1
		sheet_ranges=wb['Order Details Copy']
		sheet_ranges['P2'].value= '2018-APR-19'
		#target= wb.copy_worksheet(source)
		#wb.remove(ws)
		#wb.remove(ws2)
		#ws3=wb.active
		#ws3.title = sheetname[0]
		wb.save(filename=excelfile)