from openpyxl import load_workbook
from copy import copy, deepcopy
#from openpyxl import Workbook

class ExcelUtility(object):
	ROBOT_LIBRARY_SCOPE = 'Global'
	
	def __init__(self):
		print 'Read Cell Value in Excel File'
		
	def read_cell_value(self,excelfile,sheetname,columname,rownumber):
		wb=load_workbook(filename=excelfile,read_only=True)
		sheet_ranges=wb[sheetname]
		cellToRead=''+columname+str(rownumber)
		cellValue=sheet_ranges[cellToRead].value
		print "Cell Value: ",cellValue
		return cellValue
		
	def write_to_cell(self,excelfile,sheetname,columname,rownumber,cellvalue):
		wb=load_workbook(filename=excelfile, data_only=False)
		#sheet_ranges=wb[sheetname]
		#sheet_ranges=wb.get_sheet_by_name(sheetname)
		#ws=wb.active
		#ws = wb.get_sheet_by_name('Order Details')
		#sheetname = wb.get_sheet_names()
		#ws2 = wb.get_sheet_by_name('Cloud Connect Direct')
		#source= wb.active
		#target= wb.copy_worksheet(source)
		#wb.remove(ws)
		#wb.remove(ws2)
		#ws3=wb.active
		#ws3.title = sheetname[0]
		#for sheet in sheetname:
			#wst = wb.create_sheet(title='Order')
			#ws = wb.get_sheet_by_name(sheet)
			#source= wb.active
			#target= wb.copy_worksheet(source)
			#wb.remove(ws)
			#target.title = sheet
			#ws_copy = deepcopy(sheet)
			#ws_copy['title'] += '_' + str(i)
			#wst.append(ws_copy)

		#ws3 = wb.create_sheet(title='OrderDetails')
		#wb.active=0
		#cellToWrite=''+columname+str(rownumber)
		#sheet_ranges[cellToWrite].value= cellvalue
		#for row in sheet_ranges.iter_rows(): 
		#	for cell in row:
		#		print(cell.value)
		#sheet_ranges2=wb['Cloud Connect Direct']
		#for row in sheet_ranges2.iter_rows(): 
		#	for cell in row:
		#		print(cell.value)
		wb.save(filename=excelfile)
		#print "Cell Value: ",sheet_ranges[cellToWrite].value
		#return sheet_ranges[cellToWrite].value
		#return ws_copy