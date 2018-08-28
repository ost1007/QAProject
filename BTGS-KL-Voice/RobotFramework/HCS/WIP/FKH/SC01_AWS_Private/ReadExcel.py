from openpyxl import load_workbook

class ExcelUtility(object):
	ROBOT_LIBRARY_SCOPE = 'Global'
	
	def __init__(self):
		print 'Read Cell Value in Excel File'
		
	def read_cell_value(self,excelfile,sheetname,columname,rownumber):
		wb=load_workbook(filename=excelfile,read_only=True)
		sheet_ranges=wb[sheetname]
		cellToRead=''+columname+str(rownumber)
		cellValue=sheet_ranges[cellToRead].value
		print "Cell Value: ",cellValue
		return cellValue
		
	def write_to_cell(self,excelfile,sheetname,columname,rownumber,cellvalue):
		wb=load_workbook(filename=excelfile)
		sheet_ranges=wb[sheetname]
		cellToWrite=''+columname+str(rownumber)
		sheet_ranges[cellToWrite].value= cellvalue
		wb.save(excelfile)
		print "Cell Value: ",sheet_ranges[cellToWrite].value
		return sheet_ranges[cellToWrite].value